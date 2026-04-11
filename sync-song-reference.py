import json
import re
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import load_workbook

ROOT = Path(r'C:\Users\am186\OneDrive\ドキュメント\Project_codex')
XLSX_PATH = Path(r'C:\Users\am186\Downloads\ドラクエ楽譜曲リスト (1).xlsx')
PLAYLIST_DATA_FILE = ROOT / 'playlist-data.js'
OUTPUT_FILE = ROOT / 'song-reference-data.js'

SERIES_ORDER = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI']
SERIES_PATTERN = re.compile(r'Dragon Quest\s*(XI|X|IX|VIII|VII|VI|V|IV|III|II|I)\b', re.I)
ID_PATTERN = re.compile(r'^(XI|X|IX|VIII|VII|VI|V|IV|III|II|I)-(\d+)$', re.I)

NORMALIZATION_REPLACEMENTS = [
    ('序曲のマーチvii', '序曲のマーチ'),
    ('序曲x', '序曲'),
    ('序曲xi', '序曲'),
    ('ロトのテーマ', '序曲'),
    ('ドラゴンクエストマーチ', '序曲'),
    ('ドラゴンクエスト・マーチ', '序曲'),
    ('間奏曲', 'インテルメッツォ'),
    ('インテルメッツオ', 'インテルメッツォ'),
    ('にぎわいの街並', 'にぎわいの街並み'),
    ('木洩れ日の中で', '木もれ日の中で'),
    ('街iv', '街'),
    ('大聖堂のある街空と海と大地', '空と海と大地'),
    ('おおぞらに戦う', 'おおぞらに戦う'),
    ('急げ！ピンチだ', '急げピンチだ'),
]

REMOVE_WORDS = [
    'dragonquest', 'dragonwarrior', 'dq', 'piano', 'ドラゴンクエスト', 'ドラクエ', '交響組曲バージョン',
    '通常戦闘曲', 'テーマ', 'bgm', 'フィールド曲', '戦闘曲', 'ピアノ演奏', 'ピアノ', 'クラシック奏者',
    'クラシック勢', '演奏', '曲', 'dqii', 'dqiii', 'dqiv', 'dqv', 'dqvi', 'dqvii', 'dqviii', 'dqix', 'dqx', 'dqxi'
]

MANUAL_MATCHES = {
    ('II', 'II-6'): '遥かなる旅路',
    ('II', 'II-7'): '広野を行く',
    ('II', 'II-8'): '果てしなき世界',
    ('II', 'II-16'): 'この道わが旅',
    ('III', 'III-23'): 'そして伝説へ',
    ('IV', 'IV-2'): '間奏曲',
    ('IV', 'IV-9'): '街IV',
    ('IV', 'IV-10'): '楽しいカジノ',
    ('IV', 'IV-12'): '勇者の故郷',
    ('IV', 'IV-17'): 'エレジー',
    ('IV', 'IV-24'): '戦闘IV',
    ('IV', 'IV-26'): '悪の化身',
    ('V', 'V-2'): '間奏曲',
    ('VI', 'VI-2'): '間奏曲',
    ('VIII', 'VIII-11'): 'それ行けトーポ',
    ('VIII', 'VIII-14'): 'この想いを',
    ('VIII', 'VIII-17'): '修道僧の決意',
    ('VIII', 'VIII-18'): '急げ',
    ('VIII', 'VIII-19'): '忍び寄る影',
    ('VIII', 'VIII-23'): '詩人の世界',
    ('VIII', 'VIII-24'): '海の記憶',
    ('VIII', 'VIII-26'): '対話',
    ('VIII', 'VIII-27'): '楽しいカジノ',
    ('VIII', 'VIII-30'): '大聖堂のある街',
    ('VIII', 'VIII-34'): 'おおぞらに戦う',
    ('VIII', 'VIII-35'): '空と海と大地',
    ('IX', 'IX-3'): '天の祈り',
    ('X', 'X-7'): '夢のマイルーム',
    ('XI', 'XI-23'): '愛のこもれび',
}


def load_playlist_data():
    text = PLAYLIST_DATA_FILE.read_text(encoding='utf-8')
    text = re.sub(r'^window\.playlistData\s*=\s*', '', text)
    text = re.sub(r';\s*$', '', text)
    return json.loads(text)


def normalize_text(value: str) -> str:
    text = (value or '').lower().replace('　', ' ')
    for source, target in NORMALIZATION_REPLACEMENTS:
        text = text.replace(source, target)
    for token in ['【', '】', '[', ']', '(', ')', '/', '｜', '|', '～', '~', '・', '!', '?', '　', '、', '。', '，', '．', '…', '：', ':', '『', '』', '「', '」', '♪', '　']:
        text = text.replace(token, ' ')
    for word in REMOVE_WORDS:
        text = text.replace(word, ' ')
    text = f' {text} '
    for token in [' xi ', ' x ', ' ix ', ' viii ', ' vii ', ' vi ', ' v ', ' iv ', ' iii ', ' ii ', ' i ', ' 1 ', ' 2 ', ' 3 ', ' 4 ', ' 5 ', ' 6 ', ' 7 ', ' 8 ', ' 9 ', ' 10 ', ' 11 ']:
        text = text.replace(token, ' ')
    return ''.join(ch for ch in text if ch.isalnum() or ('ぁ' <= ch <= 'ん') or ('ァ' <= ch <= 'ン') or ('一' <= ch <= '龯'))


def confidence_for(song_title: str, track: dict) -> float:
    song_key = normalize_text(song_title)
    if not song_key:
        return 0.0
    best = 0.0
    for candidate in [track.get('title', ''), track.get('rawTitle', '')]:
        candidate_key = normalize_text(candidate)
        if not candidate_key:
            continue
        score = SequenceMatcher(None, song_key, candidate_key).ratio()
        if song_key in candidate_key or candidate_key in song_key:
            score += 0.35
        best = max(best, score)
    return best


def playlist_series_map(playlists: list[dict]) -> dict:
    mapping = {}
    for playlist in playlists:
        match = SERIES_PATTERN.search(playlist.get('title', ''))
        if match:
            mapping[match.group(1).upper()] = playlist
    return mapping


def find_track(series: str, song_id: str, song_title: str, playlist: dict | None) -> dict | None:
    if not playlist:
        return None

    tracks = playlist.get('tracks', [])
    manual_hint = MANUAL_MATCHES.get((series, song_id))
    if manual_hint:
        for track in tracks:
            haystack = f"{track.get('title', '')} {track.get('rawTitle', '')}"
            if manual_hint in haystack:
                return track

    best_track = None
    best_score = 0.0
    for track in tracks:
        score = confidence_for(song_title, track)
        if score > best_score:
            best_score = score
            best_track = track

    if best_score >= 0.72:
        return best_track

    return None


def load_song_rows():
    workbook = load_workbook(XLSX_PATH, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_by_series = {key: [] for key in SERIES_ORDER}

    for song_id, title, category in sheet.iter_rows(min_row=2, values_only=True):
        if not song_id:
            continue
        song_id = str(song_id).strip()
        title = str(title or '').strip()
        category = str(category or '').strip()
        match = ID_PATTERN.match(song_id)
        if not match:
            continue
        series_key = match.group(1).upper()
        sort_number = int(match.group(2))
        rows_by_series.setdefault(series_key, []).append({
            'id': song_id,
            'sortNumber': sort_number,
            'songTitle': title,
            'category': category,
        })

    for series_key in rows_by_series:
        rows_by_series[series_key].sort(key=lambda row: row['sortNumber'])

    return rows_by_series


def build_reference_data():
    playlists = load_playlist_data()
    playlists_by_series = playlist_series_map(playlists)
    rows_by_series = load_song_rows()
    reference = {}

    for series_key in SERIES_ORDER:
        playlist = playlists_by_series.get(series_key)
        reference_rows = []
        for row in rows_by_series.get(series_key, []):
            track = find_track(series_key, row['id'], row['songTitle'], playlist)
            reference_rows.append({
                'id': row['id'],
                'sortNumber': row['sortNumber'],
                'songTitle': row['songTitle'],
                'category': row['category'],
                'videoUrl': track.get('url') if track else '',
                'difficultyLabel': track.get('difficultyLabel') if track else '',
                'difficultyStars': track.get('difficultyStars') if track else None,
            })
        reference[series_key] = reference_rows
    return reference


def main():
    reference = build_reference_data()
    OUTPUT_FILE.write_text('window.songReferenceData = ' + json.dumps(reference, ensure_ascii=False, indent=2) + ';', encoding='utf-8')
    total = sum(len(rows) for rows in reference.values())
    linked = sum(1 for rows in reference.values() for row in rows if row['videoUrl'])
    print(f'Saved {total} song rows to {OUTPUT_FILE} ({linked} linked videos)')


if __name__ == '__main__':
    main()
